"""
Reports app views.

Provides:
  - List of all past sessions (with filtering)
  - Detailed view of a single session
  - CSV export
  - Excel export
"""

import csv
import io

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse

from attendance.models import AttendanceSession, Attendance

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@login_required
def report_list(request):
    """
    Show all past attendance sessions for the logged-in faculty.
    Supports filtering by subject name and date.
    """
    sessions = AttendanceSession.objects.filter(
        faculty=request.user
    ).prefetch_related('attendance_records').order_by('-start_time')

    # --- Filters ---
    subject_filter = request.GET.get('subject', '').strip()
    date_filter = request.GET.get('date', '').strip()

    if subject_filter:
        sessions = sessions.filter(subject_name__icontains=subject_filter)
    if date_filter:
        sessions = sessions.filter(start_time__date=date_filter)

    context = {
        'sessions': sessions,
        'subject_filter': subject_filter,
        'date_filter': date_filter,
    }
    return render(request, 'reports/report_list.html', context)


@login_required
def report_detail(request, pk):
    """Detailed view of a single session with full student list."""
    session = get_object_or_404(AttendanceSession, pk=pk, faculty=request.user)
    attendance_list = session.attendance_records.order_by('roll_number')

    # Search within this session
    query = request.GET.get('q', '').strip()
    if query:
        attendance_list = attendance_list.filter(
            roll_number__icontains=query
        ) | attendance_list.filter(full_name__icontains=query)

    context = {
        'session': session,
        'attendance_list': attendance_list,
        'total_present': session.total_present,
        'query': query,
    }
    return render(request, 'reports/report_detail.html', context)


@login_required
def export_csv(request, pk):
    """
    Export attendance data for a session as CSV.
    Returns a downloadable .csv file.
    """
    session = get_object_or_404(AttendanceSession, pk=pk, faculty=request.user)
    attendance_list = session.attendance_records.order_by('roll_number')

    filename = f"attendance_{session.subject_name.replace(' ', '_')}_{session.start_time:%Y%m%d}.csv"

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)

    # Header metadata rows
    writer.writerow(['Smart QR Attendance System'])
    writer.writerow(['Subject', session.subject_name])
    writer.writerow(['Faculty', session.faculty.get_full_name() or session.faculty.username])
    writer.writerow(['Branch', session.branch])
    writer.writerow(['Semester', session.semester])
    writer.writerow(['Section', session.section])
    writer.writerow(['Classroom', session.classroom])
    writer.writerow(['Date', session.start_time.strftime('%Y-%m-%d')])
    writer.writerow(['Duration', f'{session.duration_minutes} minutes'])
    writer.writerow(['Total Present', session.total_present])
    writer.writerow([])  # Blank separator

    # Column headers
    writer.writerow(['#', 'Roll Number', 'Full Name', 'Department', 'Semester', 'Section', 'Time'])

    for idx, record in enumerate(attendance_list, start=1):
        writer.writerow([
            idx,
            record.roll_number,
            record.full_name,
            record.department,
            record.semester,
            record.section,
            record.marked_at.strftime('%Y-%m-%d %H:%M:%S'),
        ])

    return response


@login_required
def export_excel(request, pk):
    """
    Export attendance data for a session as Excel (.xlsx).
    Uses openpyxl for formatting.
    """
    if not OPENPYXL_AVAILABLE:
        from django.contrib import messages
        messages.error(request, 'openpyxl is not installed. Please run: pip install openpyxl')
        return render(request, 'reports/report_detail.html', {
            'session': get_object_or_404(AttendanceSession, pk=pk, faculty=request.user),
        })

    session = get_object_or_404(AttendanceSession, pk=pk, faculty=request.user)
    attendance_list = session.attendance_records.order_by('roll_number')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance'

    # --- Styles ---
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    meta_font = Font(bold=True, size=10)
    center_align = Alignment(horizontal='center')

    # --- Metadata block ---
    meta_rows = [
        ['Smart QR Attendance System'],
        ['Subject:', session.subject_name],
        ['Faculty:', session.faculty.get_full_name() or session.faculty.username],
        ['Branch:', session.branch],
        ['Semester:', session.semester],
        ['Section:', session.section],
        ['Classroom:', session.classroom],
        ['Date:', session.start_time.strftime('%Y-%m-%d')],
        ['Duration:', f'{session.duration_minutes} minutes'],
        ['Total Present:', session.total_present],
        [],
    ]
    for row_data in meta_rows:
        ws.append(row_data)

    # Title row styling
    ws['A1'].font = Font(bold=True, size=14, color='1a1a2e')

    # --- Data headers ---
    col_headers = ['#', 'Roll Number', 'Full Name', 'Department', 'Semester', 'Section', 'Marked At']
    ws.append(col_headers)

    # Style header row (row 12 = 11 meta rows + 1 blank + 1 header)
    header_row_idx = len(meta_rows) + 1
    for col_idx, _ in enumerate(col_headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # --- Data rows ---
    for idx, record in enumerate(attendance_list, start=1):
        ws.append([
            idx,
            record.roll_number,
            record.full_name,
            record.department,
            record.semester,
            record.section,
            record.marked_at.strftime('%Y-%m-%d %H:%M:%S'),
        ])

    # --- Auto-fit column widths ---
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value),
            default=10
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # --- Prepare response ---
    filename = f"attendance_{session.subject_name.replace(' ', '_')}_{session.start_time:%Y%m%d}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response.write(buffer.read())
    return response
